from django.http.response import HttpResponse, HttpResponseForbidden
import openpyxl
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile
from .utils import is_admin, getMySelfId
from .models import Team, Member, TeamMember, MemberRole


def email(member):
    if not member.email_adfc:
        return member.email_private
    if not member.email_private:
        return member.email_adfc
    return member.email_adfc if member.pref else member.email_private


def ags(member):
    teamnames = [team.name for team in member.teams.all()]
    return ",".join(teamnames)


headers = [
    ("Nachname", "last_name", 20),
    ("Vorname", "first_name", 20),
    ("Geschlecht", "gender", 10),
    ("Geburtsjahr", "birthday", 12),
    ("Email-ADFC", "email_adfc", 30),
    ("Email-Privat", "email_private", 30),
    ("Email", email, 30),
    ("Telefon", "phone_primary", 20),
    ("Telefon-Alternative", "phone_secondary", 20),
    ("Postleitzahl", "address", 12),
    ("Mitgliedsnummer", "adfc_id", 17),
    ("Referenz", "reference", 50),
    ("AGs", ags, 50),
    ("Interessen", "interests", 50),
    ("Letzter Kontakt", "latest_contact", 20),
    ("Kommentar", "admin_comments", 50),
    ("Aktiv", "active", 10),
    ("Letztes Erste-Hilfe-Training", "latest_first_aid_training", 20),
    ("Registriert für Erste-Hilfe-Training",
     "registered_for_first_aid_training", 10)
]


def response(title, members, file, pref):
    if not file.endswith(".xlsx"):
        file += ".xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    ws.append([h[0] for h in headers])
    for member in members:
        member.pref = pref
        ws.append([getattr(member, h[1]) if isinstance(
            h[1], str) else h[1](member) for h in headers])
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(
            idx)].width = headers[idx-1][2]
    with NamedTemporaryFile() as tmp:
        wb.save(tmp)
        tmp.seek(0)
        content = tmp.read()
    resp = HttpResponse(
        content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = f"attachment; filename={file}"
    return resp


def excelTeam(req, teamId, file, prefEmail):
    members = []

    team = Team.objects.get(pk=teamId)
    if not is_admin(req):
        myselfId = getMySelfId(req)
        try:
            # has logged in user role="Vorsitz" in this team?
            teamMember = team.teammember_set.get(member_id=myselfId)
            role = teamMember.member_role
            if role.title != "Vorsitz":
                return HttpResponseForbidden()
        except Exception as e:
            return HttpResponseForbidden()
    return response(team.name, team.members.order_by("last_name", "first_name"), file, prefEmail == "ADFC")


def excelMembers(req, file, prefEmail):
    if not is_admin(req):
        return HttpResponseForbidden()
    members = Member.objects.all().order_by("last_name", "first_name")
    return response("Alle Aktiven", members, file, prefEmail == "ADFC")
